from email.mime import text
import re
import pandas as pd
from datetime import datetime
import os
from typing import Dict, List, Any
import argparse

class TrainingRecordExtractor:
    def __init__(self):
        """Initialize the extractor"""
        self.extracted_data = []
    
    def extract_table_data(self, table_text: str) -> Dict[str, str]:
        """Extract data from markdown table format"""
        data = {}
        
        # Split table into rows
        rows = [row.strip() for row in table_text.split('\n') if row.strip()]
        
        for row in rows:
            if '|' in row:
                # Split by pipe and clean
                cells = [cell.strip() for cell in row.split('|') if cell.strip() and cell.strip() != '---']
                
                # Process pairs of cells (key-value)
                for i in range(0, len(cells) - 1, 2):
                    if i + 1 < len(cells):
                        key = cells[i].strip()
                        value = cells[i + 1].strip()
                        if key and value:
                            data[key] = value
        
        return data
    
    def extract_review_decision(self, text: str) -> str:
        """Extract review and decision content"""
        # Simpler pattern that should work better
        pattern = r'## REVIEW AND DECISION BY TRB([\s\S]*?)(?=\|)'
        
        match = re.search(pattern, text, re.IGNORECASE)
        
        if match:
            print("Found REVIEW AND DECISION BY TRB section.")
            content = match.group(1).strip()
            # Clean up markdown formatting
            content = re.sub(r'- ~~(.+?)~~', r'- \1 (struck through)', content)
            return content
        
        print("No REVIEW AND DECISION BY TRB section found.")
        return ''
    
    def extract_vp_remarks(self, text: str) -> str:
        """Extract VP Pilots Training remarks"""
        # Look for remarks in the table - simpler pattern
        pattern = r'Remarks by VP Pilots Training[^|]*\|\s*([^|\n]+)'
        match = re.search(pattern, text, re.IGNORECASE)
        
        if match:
            return match.group(1).strip()
        
        return ''
    
    def parse_training_review_form(self, form_text: str) -> Dict[str, Any]:
        """Parse a single Training Review Board form"""
        record = {
            'Date': '',
            'Training Type': '',
            'Trainee Name': '',
            'IGA': '',
            'TRB Members': '',
            'Total Flying Experience': '',
            'Total Experience on Type': '',
            'Total PIC Experience': '',
            'REVIEW AND DECISION BY TRB': '',
            'Remarks by VP Pilots Training': ''
        }
        
        # Extract main table data (first table with Date, Training Type, etc.)
        # This table has Date, Training Type in first row and Trainee Name, IGA in second row
        main_table_pattern = r'\|\s*Date\s*\|\s*([^|]+)\s*\|\s*Training Type\s*\|\s*([^|]+)\s*\|.*?\n.*?\n\s*\|\s*Trainee Name\s*\|\s*([^|]+)\s*\|\s*IGA\s*\|\s*([^|]+)\s*\|'
        main_match = re.search(main_table_pattern, form_text, re.DOTALL | re.IGNORECASE)
        
        if main_match:
            record['Date'] = main_match.group(1).strip()
            record['Training Type'] = main_match.group(2).strip()
            record['Trainee Name'] = main_match.group(3).strip()
            record['IGA'] = main_match.group(4).strip()
        
        # Extract TRB Members from the third row of main table
        trb_pattern = r'\|\s*TRB Members\s*\|\s*([^|]+)\s*\|\s*([^|]+)\s*\|'
        trb_match = re.search(trb_pattern, form_text, re.IGNORECASE)
        if trb_match:
            trb_member1 = trb_match.group(1).strip()
            trb_member2 = trb_match.group(2).strip()
            trb_members = [member for member in [trb_member1, trb_member2] if member]
            record['TRB Members'] = ', '.join(trb_members)
        
        # Extract experience table - separate table with flying experience data
        exp_pattern = r'\|\s*Total Flying Experience\s*\|\s*Total Experience on Type\s*\|\s*Total PIC Experience\s*\|.*?\n\|\s*---.*?\n\|\s*([^|]+)\s*\|\s*([^|]+)\s*\|\s*([^|]+)\s*\|'
        exp_match = re.search(exp_pattern, form_text, re.DOTALL | re.IGNORECASE)
        
        if exp_match:
            record['Total Flying Experience'] = exp_match.group(1).strip()
            record['Total Experience on Type'] = exp_match.group(2).strip()
            record['Total PIC Experience'] = exp_match.group(3).strip()
        
        # Extract review and decision
        record['Review and Decision by TRB'] = self.extract_review_decision(form_text)
        
        # Extract VP remarks
        record['Remarks by VP Pilots Training'] = self.extract_vp_remarks(form_text)
        
        return record
    
    def extract_from_markdown(self, md_file_path: str) -> List[Dict[str, Any]]:
        """Extract all Training Review Board forms from markdown file"""
        try:
            with open(md_file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            
            # Find all Training Review Board forms
            form_pattern = r'## TRAINING REVIEW BOARD FORM(.*?)(?=\n##\s*(?!REVIEW)|Version \d+|\Z)'
            forms = re.findall(form_pattern, content, re.DOTALL | re.IGNORECASE)
            
            records = []
            for i, form_content in enumerate(forms, 1):
                print(f"Processing Training Review Board Form {i}...")
                
                # Include the heading in the form content for proper parsing
                full_form = "## TRAINING REVIEW BOARD FORM" + form_content
                record = self.parse_training_review_form(full_form)
                
                if any(record.values()):  # Only add if we extracted some data
                    records.append(record)
                    print(f"  ✓ Extracted: {record['Trainee Name']} - {record['Date']}")
                else:
                    print(f"  ✗ No data extracted from form {i}")
            
            return records
            
        except FileNotFoundError:
            raise FileNotFoundError(f"Markdown file not found: {md_file_path}")
        except Exception as e:
            raise Exception(f"Error reading markdown file: {str(e)}")
    
    def create_excel_file(self, records: List[Dict[str, Any]], output_path: str = None) -> str:
        """Create Excel file from extracted records"""
        if not records:
            raise ValueError("No records to export")
        
        # Create DataFrame
        df = pd.DataFrame(records)
        
        # Generate output filename if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"training_records_extracted_{timestamp}.xlsx"
        
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        
        # Create Excel writer with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Training Review Records', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Training Review Records']
            
            # Import openpyxl styles
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Format headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # Apply header formatting
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set minimum width of 10, maximum of 50
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Enable text wrapping for long text fields
            wrap_alignment = Alignment(wrap_text=True)
            
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, str) and len(cell.value) > 50:
                        cell.alignment = wrap_alignment
        
        return output_path
    
    def process_markdown_file(self, md_file_path: str, output_path: str = None) -> str:
        """Main processing function"""
        print(f"🚀 Starting extraction from: {md_file_path}")
        print("=" * 60)
        
        # Extract records from markdown
        records = self.extract_from_markdown(md_file_path)
        
        if not records:
            print("❌ No Training Review Board forms found in the markdown file.")
            return None
        
        print(f"\n📊 Extraction Summary:")
        print(f"  Records found: {len(records)}")
        
        # Display extracted data summary
        for i, record in enumerate(records, 1):
            print(f"\n  Record {i}:")
            print(f"    Date: '{record['Date']}'")
            print(f"    Training Type: '{record['Training Type']}'")
            print(f"    Trainee Name: '{record['Trainee Name']}'")
            print(f"    IGA: '{record['IGA']}'")
            print(f"    TRB Members: '{record['TRB Members']}'")
            print(f"    Total Flying Experience: '{record['Total Flying Experience']}'")
            print(f"    Total Experience on Type: '{record['Total Experience on Type']}'")
            print(f"    Total PIC Experience: '{record['Total PIC Experience']}'")
            print(f"    Review and Decision by TRB: '{record['Review and Decision by TRB']}'")
            print(f"    Remarks by VP Pilots Training: '{record['Remarks by VP Pilots Training']}'")
        
        # Create Excel file
        output_file = self.create_excel_file(records, output_path)
        
        print(f"\n✅ Excel file created successfully:")
        print(f"   📁 File: {output_file}")
        print(f"   📊 Records: {len(records)}")
        print(f"   📋 Columns: {len(records[0].keys())}")
        
        return output_file

def main():
    """Main function to run the extractor"""
    parser = argparse.ArgumentParser(description='Extract Training Review Board data from markdown to Excel')
    parser.add_argument('input_file', help='Path to the markdown file')
    parser.add_argument('-o', '--output', help='Output Excel file path (optional)')
    
    args = parser.parse_args()
    
    try:
        # Initialize extractor
        extractor = TrainingRecordExtractor()
        
        # Validate input file
        if not os.path.exists(args.input_file):
            print(f"❌ Error: Input file not found: {args.input_file}")
            return 1
        
        if not args.input_file.lower().endswith('.md'):
            print(f"⚠️  Warning: Input file doesn't have .md extension")
        
        # Process the file
        output_file = extractor.process_markdown_file(args.input_file, args.output)
        
        if output_file:
            print(f"\n🎉 Processing completed successfully!")
            print(f"📁 Output file: {output_file}")
        else:
            print("❌ No data was extracted.")
            return 1
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return 1
    
    return 0

if __name__ == "__main__":
    # If no command line arguments, run in interactive mode
    if len(os.sys.argv) == 1:
        print("🔍 Interactive mode")
        print("=" * 40)
        
        # Get input file
        input_file = input("Enter path to markdown file: ").strip()
        if input_file.startswith('"') and input_file.endswith('"'):
            input_file = input_file[1:-1]  # Remove quotes
        
        # Get output file (optional)
        output_file = input("Enter output Excel file path (press Enter for auto-generated): ").strip()
        if not output_file:
            output_file = None
        
        try:
            extractor = TrainingRecordExtractor()
            result = extractor.process_markdown_file(input_file, output_file)
            if result:
                print(f"\n✅ Success! Check: {result}")
        except Exception as e:
            print(f"❌ Error: {str(e)}")
    else:
        # Command line mode
        exit(main())